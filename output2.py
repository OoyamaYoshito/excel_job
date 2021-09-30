# coding: UTF-8

import sys
import openpyxl
import numpy as np
import pandas as pd
from PIL import Image
import calculate
import chart_python
import matplotlib.font_manager

#一人分のアンケート結果を出力
def question_result_output(personal_info, answerdata, ws, labels, outputpath='.'):

    #個人情報挿入
    for i, data in enumerate(personal_info):
        if i < 1:
            ws.cell(4,1+i,data)
        else:
            ws.cell(4,2+i,data)
    #ws.cell(4,1,personal_info)
    
    #数字データ結果挿入
    for i, yeardata in enumerate(answerdata):
        ws.cell(7+i,1,labels[i])
        for j,data in enumerate(yeardata):
            ws.cell(7+i,2+j,data)

    #グラフ画像挿入
    chart_python.draw_chart(answerdata, labels, outputpath)
    img = openpyxl.drawing.image.Image(outputpath + '/graph.png')
    ws.add_image( img, 'B13' )

#一クラス分のExcel生成
def output_students(students, outputpath='.', studentpath='studentlist', answerpath='answersdata'):
    df_answerdata = calculate.get_answerdata(answerpath)
    df_classdata = calculate.get_studentdata(students, studentpath)
    df_combineds, df_mean = calculate.calc_mean(df_answerdata, df_classdata)
    classdata, answerdatas, averagedata = df_classdata, df_combineds, df_mean

    fn = outputpath + '/output.xlsx'
    wb = openpyxl.load_workbook('templete.xlsx')
    ws = wb.worksheets[0]
    if classdata.index.size == 0:
        print ('11クラスのデータが取得できませんでした')
        raise ValueError('11クラスのデータが取得できませんでした')
    for sheet_name in list(classdata.index):
        ws_copy = wb.copy_worksheet(ws)
        ws_copy.title=str(sheet_name)
    try:
        wb.save(fn)
    except PermissionError:
        print ('13出力先のファイルが開かれているため、出力できません: ' + fn)
        raise ValueError('13出力先のファイルが開かれているため、出力できません: ' + fn)
    for i, student_number in enumerate(list(classdata.index)):
        wb = openpyxl.load_workbook(fn)
        personal_info = classdata.loc[student_number].values
        stgrade = personal_info[3][0]
        stclass = personal_info[3][1]
        personal_info[0] = str(student_number) + " " + personal_info[0]
        answerdata = []
        labels = ()
        for (ym,yeardata) in answerdatas:
            if student_number in list(yeardata.index):
                answerdata.append(yeardata.loc[student_number].values)
                labels = labels + (ym,)
        if len(answerdata) > 3:
            print ('14回答回数が3回を超えている学生がいます: ' + str(student_number))
            raise ValueError('14回答回数が3回を超えている学生がいます: ' + str(student_number))
        #if stclass in ["G","H","I"]:
        #    labels = ('1年前期終了時', '2年後期開始時', '2年後期終了時')
        #else:
        #    labels = ('1年前期終了時', '2年前期開始時', '2年前期終了時')
        for _ in range(3-len(labels)):
            labels = labels + ('',)
        labels = labels + (labels[len(answerdata)-1]+'平均',)
        for _ in range(3-len(answerdata)):
            answerdata.append([None,None,None,None,None,None,None,None])
        answerdata.append(averagedata)
        question_result_output(personal_info,answerdata,wb.worksheets[i+1],labels,outputpath)
        wb.save(fn)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print ("Usage: python "+sys.argv[0]+" <学籍番号[,学籍番号[,学籍番号[...]]]>")
        sys.exit()
    else:
        args = sys.argv
        students = []
        for i in range(1,len(args)):
            st = args[i].split(',')
            students.extend(st)
        print(students)
        output_students(students)
